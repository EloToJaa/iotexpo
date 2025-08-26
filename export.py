import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import re

STATE_PATH = "cp.json"


def read_state():
    with open(STATE_PATH, "r", encoding="utf-8") as file:
        data = json.load(file)
    return data


def remove_illegal(input: str) -> str:
    return re.sub(r"[\x00-\x1F]", "", input)


def run_export(
    workbook: openpyxl.Workbook,
    n: int,
    xlsx_data: list[list[str]],
    header_row: list[str],
):
    sheet = workbook.create_sheet(title=f"Exhibitors IoT Expo China Hall {n}")

    sheet.append(header_row)
    for row in xlsx_data:
        new_row = [remove_illegal(cell) for cell in row]
        sheet.append(new_row)

    table_range = f"A1:F{len(xlsx_data)}"

    table = Table(displayName="ExhibitorTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style

    sheet.add_table(table)

    for row in range(2, len(xlsx_data) + 1):
        cell = sheet.cell(row=row, column=5)
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    xlsx_data = [header_row]


state = read_state()
header_row = [
    "Company Name",
    "Exhibitor Name",
    "Booth Number",
    "Company Address",
    "Logo URL",
    "Company Introduction",
]

workbook = openpyxl.Workbook()

for i in range(9, 13):
    run_export(workbook, i, state[str(i)]["data"], header_row)

workbook.save("exhibitors.xlsx")
